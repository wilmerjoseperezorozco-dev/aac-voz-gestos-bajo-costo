"""Genera la bitácora maestra en Excel a partir de TODOS los registros ya
existentes (registros/*.csv, reportes/validacion*.json) — no se captura
nada dos veces, este script solo consolida y da formato.

Se guarda en registros/bitacora_maestra.xlsx — excluida de git por el
mismo patrón que protege el resto de registros/ (datos operativos, no
para el repositorio público).

Uso (rodar después de cada sesión para mantenerla al día):
    python src/generar_bitacora.py
"""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

RAIZ = Path(__file__).resolve().parent.parent
DIR_REGISTROS = RAIZ / "registros"
DIR_REPORTES = RAIZ / "reportes"

AZUL = "1F4E79"
GRIS_CLARO = "F2F2F2"
FUENTE = "Calibri"


def leer_csv(ruta: Path) -> list[dict]:
    if not ruta.exists():
        return []
    with ruta.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def escribir_hoja(wb: Workbook, nombre: str, columnas: list[str], filas: list[dict]) -> None:
    hoja = wb.create_sheet(nombre)
    for c, encabezado in enumerate(columnas, start=1):
        celda = hoja.cell(row=1, column=c, value=encabezado)
        celda.font = Font(name=FUENTE, bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, fila in enumerate(sorted(filas, key=lambda x: x.get("fecha_hora", "")), start=2):
        for c, col in enumerate(columnas, start=1):
            valor = fila.get(col, "")
            celda = hoja.cell(row=r, column=c, value=valor)
            celda.font = Font(name=FUENTE, size=10)
            if r % 2 == 0:
                celda.fill = PatternFill("solid", fgColor=GRIS_CLARO)

    for c, encabezado in enumerate(columnas, start=1):
        ancho = max(14, min(45, len(encabezado) + 4))
        hoja.column_dimensions[get_column_letter(c)].width = ancho
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}{max(2, len(filas) + 1)}"


def hoja_fichas_sesion(wb: Workbook) -> None:
    columnas = ["fecha_hora", "mic_fijo", "ruido_apagado", "fondo_liso",
                "luz_frontal", "marcas_piso", "evaluador_ciego",
                "celular_lateral", "presentes", "animo_yp", "notas"]
    filas = leer_csv(DIR_REGISTROS / "fichas_sesion.csv")
    escribir_hoja(wb, "Fichas de sesion", columnas, filas)


def hoja_sesiones_grabacion(wb: Workbook) -> None:
    columnas = ["fecha_hora", "canal", "alias", "palabra_o_gesto",
                "archivo", "observaciones"]
    filas = []
    for f in leer_csv(DIR_REGISTROS / "sesiones.csv"):
        filas.append({"fecha_hora": f["fecha_hora"], "canal": "voz",
                      "alias": f["alias"], "palabra_o_gesto": f["palabra"],
                      "archivo": f["archivo"], "observaciones": f.get("observaciones", "")})
    for f in leer_csv(DIR_REGISTROS / "sesiones_gestos.csv"):
        filas.append({"fecha_hora": f["fecha_hora"], "canal": "gesto",
                      "alias": f["alias"], "palabra_o_gesto": f["gesto"],
                      "archivo": f["archivo"], "observaciones": f.get("observaciones", "")})
    for f in leer_csv(DIR_REGISTROS / "sesiones_multimodal.csv"):
        filas.append({"fecha_hora": f["fecha_hora"], "canal": "multimodal",
                      "alias": f["alias"], "palabra_o_gesto": f["significado"],
                      "archivo": f"{f['archivo_audio']} + {f['archivo_gesto']}",
                      "observaciones": ""})
    escribir_hoja(wb, "Sesiones de grabacion", columnas, filas)


def hoja_predicciones_voz(wb: Workbook) -> None:
    columnas = ["fecha_hora", "archivo_origen", "evaluador_ciego",
                "prediccion", "confianza", "correcta"]
    filas = []
    for ruta in sorted(DIR_REGISTROS.glob("predicciones*.csv")):
        if "gestos" in ruta.stem or "multimodal" in ruta.stem:
            continue
        for f in leer_csv(ruta):
            filas.append({
                "fecha_hora": f.get("fecha_hora", ""),
                "archivo_origen": ruta.name,
                "evaluador_ciego": f.get("prediccion_evaluador_ciego", ""),
                "prediccion": f.get("prediccion", ""),
                "confianza": f.get("confianza", ""),
                "correcta": f.get("correcta", ""),
            })
    escribir_hoja(wb, "Predicciones voz", columnas, filas)


def hoja_predicciones_gestos(wb: Workbook) -> None:
    columnas = ["fecha_hora", "archivo_origen", "evaluador_ciego",
                "gesto", "significado", "confianza", "correcta"]
    filas = []
    for ruta in sorted(DIR_REGISTROS.glob("predicciones_gestos*.csv")):
        for f in leer_csv(ruta):
            filas.append({
                "fecha_hora": f.get("fecha_hora", ""),
                "archivo_origen": ruta.name,
                "evaluador_ciego": f.get("significado_evaluador_ciego", ""),
                "gesto": f.get("gesto", ""),
                "significado": f.get("significado", ""),
                "confianza": f.get("confianza", ""),
                "correcta": f.get("correcta", ""),
            })
    for ruta in sorted(DIR_REGISTROS.glob("predicciones_multimodal*.csv")):
        for f in leer_csv(ruta):
            filas.append({
                "fecha_hora": f.get("fecha_hora", ""),
                "archivo_origen": ruta.name,
                "evaluador_ciego": f.get("prediccion_evaluador_ciego", ""),
                "gesto": f.get("modo", ""),
                "significado": f.get("resultado_fusion", ""),
                "confianza": f.get("confianza_voz", ""),
                "correcta": f.get("correcta", ""),
            })
    escribir_hoja(wb, "Predicciones gestos-multimodal", columnas, filas)


def hoja_descartes(wb: Workbook) -> None:
    columnas = ["fecha_hora", "palabra", "archivo", "rms", "motivo"]
    escribir_hoja(wb, "Descartes", columnas, leer_csv(DIR_REGISTROS / "descartes.csv"))


def hoja_tablero_nucleo(wb: Workbook) -> None:
    columnas = ["fecha_hora", "simbolos_seleccionados", "contexto",
                "oracion_generada", "confirmada"]
    escribir_hoja(wb, "Tablero nucleo (Capa 1+2)", columnas,
                  leer_csv(DIR_REGISTROS / "predicciones_tablero.csv"))


def hoja_validaciones(wb: Workbook) -> None:
    columnas = ["fecha_hora", "canal", "n_muestras", "n_clases",
                "exactitud_global_pct", "archivo"]
    filas = []
    for ruta in sorted(DIR_REPORTES.glob("validacion*.json")):
        try:
            d = json.loads(ruta.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            continue
        canal = "gestos" if "gestos" in ruta.stem else "voz"
        filas.append({
            "fecha_hora": d.get("fecha", ruta.stem),
            "canal": canal,
            "n_muestras": d.get("total_muestras", ""),
            "n_clases": len(d.get("palabras", [])),
            "exactitud_global_pct": round(d.get("exactitud_global", 0) * 100, 1),
            "archivo": ruta.name,
        })
    escribir_hoja(wb, "Validaciones LOOCV", columnas, filas)


def hoja_resumen(wb: Workbook) -> None:
    hoja = wb.create_sheet("Resumen", 0)
    hoja.sheet_view.showGridLines = False
    hoja["A1"] = "Bitácora maestra — MVP predicción de voz y gestos (YP)"
    hoja["A1"].font = Font(name=FUENTE, bold=True, size=16, color=AZUL)
    hoja["A2"] = "Generada automáticamente a partir de registros/ y reportes/ — volver a correr src/generar_bitacora.py tras cada sesión."
    hoja["A2"].font = Font(name=FUENTE, italic=True, size=10, color="595959")

    etiquetas_formulas = [
        ("Total fichas de sesión", "=COUNTA('Fichas de sesion'!A:A)-1"),
        ("Total grabaciones (voz+gesto+multimodal)", "=COUNTA('Sesiones de grabacion'!A:A)-1"),
        ("Total intentos de predicción en vivo (voz)", "=COUNTA('Predicciones voz'!A:A)-1"),
        ("  — confirmados correctos (voz)", '=COUNTIF(\'Predicciones voz\'!F:F,"si")'),
        ("  — confirmados incorrectos (voz)", '=COUNTIF(\'Predicciones voz\'!F:F,"no")'),
        ("Total intentos gestos/multimodal", "=COUNTA('Predicciones gestos-multimodal'!A:A)-1"),
        ("Total intentos tablero núcleo (Capa 1+2)", "=COUNTA('Tablero nucleo (Capa 1+2)'!A:A)-1"),
        ("  — oraciones confirmadas correctas", '=COUNTIF(\'Tablero nucleo (Capa 1+2)\'!E:E,"si")'),
        ("Total muestras descartadas", "=COUNTA(Descartes!A:A)-1"),
        ("Total entrenamientos (LOOCV) registrados", "=COUNTA('Validaciones LOOCV'!A:A)-1"),
    ]
    fila = 4
    for etiqueta, formula in etiquetas_formulas:
        hoja.cell(row=fila, column=1, value=etiqueta).font = Font(name=FUENTE, size=11)
        celda_valor = hoja.cell(row=fila, column=2, value=formula)
        celda_valor.font = Font(name=FUENTE, size=11, bold=True, color=AZUL)
        fila += 1

    hoja.column_dimensions["A"].width = 45
    hoja.column_dimensions["B"].width = 16


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # quitar la hoja "Sheet" en blanco por defecto

    hoja_fichas_sesion(wb)
    hoja_sesiones_grabacion(wb)
    hoja_predicciones_voz(wb)
    hoja_predicciones_gestos(wb)
    hoja_tablero_nucleo(wb)
    hoja_descartes(wb)
    hoja_validaciones(wb)
    hoja_resumen(wb)  # al final para que el Resumen quede como primera hoja (index 0)

    DIR_REGISTROS.mkdir(exist_ok=True)
    ruta_salida = DIR_REGISTROS / "bitacora_maestra.xlsx"
    wb.save(ruta_salida)
    print(f"Bitácora generada: {ruta_salida.relative_to(RAIZ)}")
    print("Recuerda: este archivo NO se sube a GitHub (registros/ esta en .gitignore).")


if __name__ == "__main__":
    main()
